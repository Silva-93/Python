# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# Criando um workbook
workbook = Workbook()
# worksheet: Worksheet = workbook.active

# Criando os cabeçalhos
#          1°linha  1° coluna
# worksheet.cell(1, 1, 'Nome')

# #          1°linha  2° coluna
# worksheet.cell(1, 2, 'Idade')

# #          1°linha  3° coluna
# worksheet.cell(1, 3, 'Nota')


# Nomeando a planilha
sheet_name = 'Minha Planilha'

# Criando a planilha
workbook.create_sheet(sheet_name, 0)

# Selecionando a planilha
worksheet: Worksheet = workbook[sheet_name]

# Removendo uma planilha
workbook.remove(workbook['Sheet'])  # nome da planilha

students = [
    # nome      idade       nota
    ['João',     14,        5.5 ],
    ['Maria',    13,        9.7 ],
    ['Luiz',     15,        8.8 ],
    ['Alberto',  16,        10  ],
]

# for i, students_row in enumerate(students, start=2):
#     for j, students_column in enumerate(students_row, start=1):
#         worksheet.cell(i, j, students_column)

# Outra forma

for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)
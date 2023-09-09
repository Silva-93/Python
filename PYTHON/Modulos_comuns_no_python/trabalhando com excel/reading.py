# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell 
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# Carregando um arquivo do excel
workbook: Workbook = load_workbook(WORKBOOK_PATH)
# worksheet: Worksheet = workbook.active

# Nomeando a planilha
sheet_name = 'Minha Planilha'

# Selecionando a planilha
worksheet: Worksheet = workbook[sheet_name]

row: tuple[Cell]
for row in worksheet.iter_rows():
    for cell in row:
        print(cell.value, end='\t')

        # Alterando dados de uma célula
        if cell.value == 'Maria':
            #            N° linha  N° col  idade alterada
            worksheet.cell(cell.row, 2, 20)

    print()

# Alterando dados de uma célula
# worksheet['B2'].value = 15

workbook.save(WORKBOOK_PATH)